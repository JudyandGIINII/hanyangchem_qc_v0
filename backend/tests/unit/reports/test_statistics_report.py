from __future__ import annotations

import io
import os
from collections.abc import Generator
from datetime import UTC, date, datetime
from decimal import Decimal
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from hyc_api.reports.deterministic import workbook_digest
from hyc_api.reports.statistics import (
    StatisticsRow,
    build_statistics_rows,
    calculate_quality_statistics_data,
    query_approved_inspection_cases,
    render_supplier_quality_statistics_report,
)
from hyc_data.models import (
    Base,
    DecisionSnapshotRow,
    Document,
    DocumentAllocationLink,
    DocumentSection,
    ExtractionRun,
    InboundReceipt,
    InspectionCase,
    Material,
    MaterialLot,
    ReceiptLotAllocation,
    SpecProfile,
    SpecVersion,
    Supplier,
)

POSTGRES_DSN = os.environ.get("HYC_P3_TEST_POSTGRES_DSN") or os.environ.get(
    "HYC_P2_TEST_POSTGRES_DSN"
)
# Only the two database-backed cases below carry the postgres marker.  Marking the
# whole module would deselect the in-memory regressions -- the Asia/Seoul month
# boundary and the all-cells-are-strings rule -- from `make check`, which is the
# gate that actually runs on every change.


@pytest.fixture
def session() -> Generator[Session, None, None]:
    if POSTGRES_DSN:
        engine = create_engine(POSTGRES_DSN)
    else:
        engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as s:
        yield s
    engine.dispose()


def test_rule_1_and_rule_2_query_helper_exclusions(session: Session) -> None:
    supplier = Supplier(id=uuid4(), name="세계로비드")
    material = Material(id=uuid4(), name="염화칼슘")
    session.add_all([supplier, material])
    session.flush()

    lot = MaterialLot(
        id=uuid4(),
        supplier_id=supplier.id,
        material_id=material.id,
        identity_policy_version="v1",
        identity_key="LOT-RULE-001",
        identity_status="CANONICAL",
    )
    session.add(lot)
    session.flush()

    receipt = InboundReceipt(
        id=uuid4(),
        inbound_no="INB-RULE-1",
        supplier_id=supplier.id,
        receipt_date=date(2026, 5, 1),
    )
    session.add(receipt)
    session.flush()

    alloc = ReceiptLotAllocation(
        id=uuid4(),
        inbound_receipt_id=receipt.id,
        material_lot_id=lot.id,
        quantity=Decimal("10.00"),
        quantity_unit="kg",
    )
    session.add(alloc)
    session.flush()

    # Case 1: Approved case (has DecisionSnapshotRow, status ACCEPTED) -> MUST BE INCLUDED
    c1 = InspectionCase(
        id=uuid4(),
        receipt_lot_allocation_id=alloc.id,
        spec_version_id=uuid4(),
        status="ACCEPTED",
        final_decision=None,
    )
    # Case 2: Unapproved case (NO DecisionSnapshotRow) -> MUST BE EXCLUDED (Rule 1)
    c2 = InspectionCase(
        id=uuid4(),
        receipt_lot_allocation_id=alloc.id,
        spec_version_id=uuid4(),
        status="DRAFT",
        final_decision=None,
    )
    # Case 3: Cancelled case (has DecisionSnapshotRow, status CANCELLED) -> EXCLUDED (Rule 2)
    c3 = InspectionCase(
        id=uuid4(),
        receipt_lot_allocation_id=alloc.id,
        spec_version_id=uuid4(),
        status="CANCELLED",
        final_decision=None,
    )
    session.add_all([c1, c2, c3])
    session.flush()

    c1.final_decision = "ACCEPTED"
    c2.final_decision = "ACCEPTED"
    c3.final_decision = "REJECTED"
    session.flush()

    payload_s1 = {
        "overall_decision": "ACCEPTED",
        "spec_version": {"semantic_version": "1.0.0", "status": "ACTIVE"},
        "approver": {"actor_id": str(uuid4()), "role": "LEAD"},
    }
    payload_s3 = {
        "overall_decision": "REJECTED",
        "spec_version": {"semantic_version": "1.0.0", "status": "ACTIVE"},
        "approver": {"actor_id": str(uuid4()), "role": "LEAD"},
    }

    s1 = DecisionSnapshotRow(
        id=uuid4(),
        inspection_case_id=c1.id,
        payload=payload_s1,
        content_hash="1" * 64,
        created_at=datetime(2026, 5, 10, 10, 0, 0, tzinfo=UTC),
    )
    s3 = DecisionSnapshotRow(
        id=uuid4(),
        inspection_case_id=c3.id,
        payload=payload_s3,
        content_hash="3" * 64,
        created_at=datetime(2026, 5, 10, 10, 0, 0, tzinfo=UTC),
    )

    session.add_all([s1, s3])
    session.commit()

    records = query_approved_inspection_cases(session)
    fetched_case_ids = [c.id for c, _ in records]
    assert c1.id in fetched_case_ids
    assert c2.id not in fetched_case_ids  # Rule 1: no snapshot row
    assert c3.id not in fetched_case_ids  # Rule 2: CANCELLED status


def test_timezone_month_boundary_regression_in_memory() -> None:
    """Pin the choice that period bucketing is computed in Asia/Seoul calendar.

    Timestamp: 2026-04-30 20:00:00 UTC
    In UTC calendar: April 2026 (2026-04-30)
    In KST calendar (+9h): May 1, 2026 05:00:00 KST (2026-05-01)
    """
    snapshot_time = datetime(2026, 4, 30, 20, 0, 0, tzinfo=UTC)
    row = StatisticsRow(
        inspection_case_id="case-tz-1",
        snapshot_created_at=snapshot_time,
        status="ACCEPTED",
        final_decision="ACCEPTED",
        material_code="MAT-001",
        material_name="염화칼슘",
    )
    observed = datetime(2026, 8, 12, 0, 0, tzinfo=UTC)

    # May 2026 KST report (2026-05-01 to 2026-05-31) MUST include this record!
    may_bytes = render_supplier_quality_statistics_report(
        [row], period_start="2026-05-01", period_end="2026-05-31", observed_at=observed
    )
    may_wb = load_workbook(io.BytesIO(may_bytes))
    may_summary = list(may_wb["품질통계요약"].iter_rows(values_only=True))
    total_in_may = next(r[1] for r in may_summary if r[0] == "총 입고/검사 건수")
    assert total_in_may == "1"

    # April 2026 KST report (2026-04-01 to 2026-04-30) MUST EXCLUDE this record!
    apr_bytes = render_supplier_quality_statistics_report(
        [row], period_start="2026-04-01", period_end="2026-04-30", observed_at=observed
    )
    apr_wb = load_workbook(io.BytesIO(apr_bytes))
    apr_summary = list(apr_wb["품질통계요약"].iter_rows(values_only=True))
    total_in_apr = next(r[1] for r in apr_summary if r[0] == "총 입고/검사 건수")
    assert total_in_apr == "0"


def test_statistics_report_in_memory_metrics_and_all_cell_values_are_strings() -> None:
    row1 = StatisticsRow(
        inspection_case_id="case-1",
        snapshot_created_at=datetime(2026, 5, 2, 12, 0, 0, tzinfo=UTC),
        created_at=datetime(2026, 5, 1, 0, 0, 0, tzinfo=UTC),
        status="REJECTED",
        final_decision="REJECTED",
        material_code="MAT-001",
        material_name="염화칼슘",
        has_coa_document=False,
        ocr_review_required=True,
        internal_test_required=True,
        internal_test_completed=True,
        failed_test_items=(("ITEM-01", "수분함량"),),
        nonconformance_dispositions=(("RETURN", "반품"),),
    )
    row2 = StatisticsRow(
        inspection_case_id="case-2",
        snapshot_created_at=datetime(2026, 5, 5, 12, 0, 0, tzinfo=UTC),
        created_at=datetime(2026, 5, 4, 0, 0, 0, tzinfo=UTC),
        status="ACCEPTED",
        final_decision="ACCEPTED",
        material_code="MAT-001",
        material_name="염화칼슘",
        has_coa_document=True,
        ocr_review_required=False,
        internal_test_required=False,
        internal_test_completed=False,
    )
    rows = [row1, row2]
    observed = datetime(2026, 8, 12, 0, 0, tzinfo=UTC)

    report_bytes = render_supplier_quality_statistics_report(
        rows, period_start="2026-05-01", period_end="2026-05-31", observed_at=observed
    )
    assert len(report_bytes) > 0
    assert workbook_digest(report_bytes) == workbook_digest(report_bytes)

    wb = load_workbook(io.BytesIO(report_bytes))
    assert list(wb.sheetnames) == [
        "품질통계요약",
        "품목별부적합",
        "검사항목별부적합",
        "부적합처리방안",
    ]

    summary = list(wb["품질통계요약"].iter_rows(values_only=True))
    metrics_map = {str(r[0]): str(r[1]) for r in summary[1:] if len(r) >= 2}

    assert metrics_map["총 입고/검사 건수"] == "2"
    assert metrics_map["적합 건수"] == "1"
    assert metrics_map["부적합 건수"] == "1"
    assert metrics_map["부적합률"] == "50.00%"
    assert metrics_map["COA 누락률"] == "50.00%"
    assert metrics_map["OCR 검토 필요율"] == "50.00%"

    for name in wb.sheetnames:
        sheet = wb[name]
        for r in sheet.iter_rows(values_only=True):
            for cell_val in r:
                if cell_val is not None:
                    err_msg = f"Cell in {name} is not str: {type(cell_val)}"
                    assert isinstance(cell_val, str), err_msg


def test_ocr_review_rate_is_reported_as_percentage() -> None:
    """Verify that ocr_review_rate is computed as a Decimal percentage string."""
    data = calculate_quality_statistics_data(
        [
            {
                "inspection_case_id": "c1",
                "snapshot_created_at": datetime(2026, 8, 10, 3, 0, tzinfo=UTC),
                "status": "ACCEPTED",
                "final_decision": "ACCEPTED",
                "material_code": "M1",
                "material_name": "염화칼슘",
                "supplier_code": "S1",
                "supplier_name": "세계로비드",
                "ocr_review_required": True,
            },
            {
                "inspection_case_id": "c2",
                "snapshot_created_at": datetime(2026, 8, 11, 3, 0, tzinfo=UTC),
                "status": "ACCEPTED",
                "final_decision": "ACCEPTED",
                "material_code": "M1",
                "material_name": "염화칼슘",
                "supplier_code": "S1",
                "supplier_name": "세계로비드",
                "ocr_review_required": False,
            },
        ],
        "2026-08-01",
        "2026-08-31",
        datetime(2026, 9, 1, tzinfo=UTC),
    )
    assert data["ocr_review_rate"] == "50.00%"


def test_ocr_review_rate_derived_from_extraction_run_both_directions(session: Session) -> None:
    """Regression test proving OCR review rate is genuinely non-zero for a case with an
    ExtractionRun and zero for a case without.

    Proves both directions via build_statistics_rows against the real schema.
    """
    supplier = Supplier(id=uuid4(), name="세계로비드", supplier_code="SUP-001")
    material = Material(id=uuid4(), name="염화칼슘", material_code="MAT-001")
    session.add_all([supplier, material])
    session.flush()

    spec_prof = SpecProfile(
        id=uuid4(),
        material_id=material.id,
        name="염화칼슘 규격",
    )
    session.add(spec_prof)
    session.flush()

    spec_ver = SpecVersion(
        id=uuid4(),
        spec_profile_id=spec_prof.id,
        version=1,
        status="ACTIVE",
        effective_from=date(2026, 1, 1),
    )
    session.add(spec_ver)
    session.flush()

    # Lot 1 & Receipt 1 & Alloc 1: WITH ExtractionRun
    lot1 = MaterialLot(
        id=uuid4(),
        supplier_id=supplier.id,
        material_id=material.id,
        identity_policy_version="v1",
        identity_key="LOT-OCR-001",
        identity_status="CANONICAL",
    )
    receipt1 = InboundReceipt(
        id=uuid4(),
        inbound_no="INB-OCR-1",
        supplier_id=supplier.id,
        receipt_date=date(2026, 5, 1),
    )
    session.add_all([lot1, receipt1])
    session.flush()

    alloc1 = ReceiptLotAllocation(
        id=uuid4(),
        inbound_receipt_id=receipt1.id,
        material_lot_id=lot1.id,
        quantity=Decimal("10.00"),
        quantity_unit="kg",
    )
    session.add(alloc1)
    session.flush()

    # Document & Section & AllocationLink & ExtractionRun for Case 1
    doc1 = Document(
        id=uuid4(),
        document_type="COA",
        checksum_sha256="1" * 64,
        original_filename="coa1.pdf",
        storage_key="storage/coa1.pdf",
        immutable=True,
    )
    session.add(doc1)
    session.flush()

    sec1 = DocumentSection(
        id=uuid4(),
        document_id=doc1.id,
        section_index=0,
        page_from=1,
        page_to=1,
        status="MATCHED",
    )
    session.add(sec1)
    session.flush()

    link1 = DocumentAllocationLink(
        id=uuid4(),
        document_section_id=sec1.id,
        receipt_lot_allocation_id=alloc1.id,
        match_status="CONFIRMED",
    )
    session.add(link1)
    session.flush()


    run1 = ExtractionRun(
        id=uuid4(),
        document_id=doc1.id,
        provider_name="LOCAL_PADDLE_OCR",
        status="REVIEW_REQUIRED",
        candidate_payload={"items": []},
    )
    session.add(run1)
    session.flush()

    c1 = InspectionCase(
        id=uuid4(),
        receipt_lot_allocation_id=alloc1.id,
        spec_version_id=spec_ver.id,
        status="ACCEPTED",
        final_decision=None,
        created_at=datetime(2026, 5, 1, 0, 0, tzinfo=UTC),
    )
    session.add(c1)
    session.flush()

    c1.final_decision = "ACCEPTED"
    session.flush()

    s1 = DecisionSnapshotRow(
        id=uuid4(),
        inspection_case_id=c1.id,
        payload={
            "overall_decision": "ACCEPTED",
            "spec_version": {"semantic_version": "1.0.0", "status": "ACTIVE"},
            "approver": {"actor_id": str(uuid4()), "role": "LEAD"},
        },
        content_hash="1" * 64,
        created_at=datetime(2026, 5, 2, 12, 0, tzinfo=UTC),
    )
    session.add(s1)
    session.flush()

    # Direction 1: Single case with ExtractionRun -> ocr_review_required is True,
    # ocr_review_rate is "100.00%"
    rows_c1 = build_statistics_rows(session)
    assert len(rows_c1) == 1
    assert rows_c1[0].ocr_review_required is True

    data_c1 = calculate_quality_statistics_data(
        rows_c1, "2026-05-01", "2026-05-31", datetime(2026, 5, 3, tzinfo=UTC)
    )
    assert data_c1["ocr_review_rate"] == "100.00%"

    # Lot 2 & Receipt 2 & Alloc 2: WITHOUT ExtractionRun
    lot2 = MaterialLot(
        id=uuid4(),
        supplier_id=supplier.id,
        material_id=material.id,
        identity_policy_version="v1",
        identity_key="LOT-OCR-002",
        identity_status="CANONICAL",
    )
    receipt2 = InboundReceipt(
        id=uuid4(),
        inbound_no="INB-OCR-2",
        supplier_id=supplier.id,
        receipt_date=date(2026, 5, 1),
    )
    session.add_all([lot2, receipt2])
    session.flush()

    alloc2 = ReceiptLotAllocation(
        id=uuid4(),
        inbound_receipt_id=receipt2.id,
        material_lot_id=lot2.id,
        quantity=Decimal("10.00"),
        quantity_unit="kg",
    )
    session.add(alloc2)
    session.flush()

    c2 = InspectionCase(
        id=uuid4(),
        receipt_lot_allocation_id=alloc2.id,
        spec_version_id=spec_ver.id,
        status="ACCEPTED",
        final_decision=None,
        created_at=datetime(2026, 5, 3, 0, 0, tzinfo=UTC),
    )
    session.add(c2)
    session.flush()

    c2.final_decision = "ACCEPTED"
    session.flush()

    s2 = DecisionSnapshotRow(
        id=uuid4(),
        inspection_case_id=c2.id,
        payload={
            "overall_decision": "ACCEPTED",
            "spec_version": {"semantic_version": "1.0.0", "status": "ACTIVE"},
            "approver": {"actor_id": str(uuid4()), "role": "LEAD"},
        },
        content_hash="2" * 64,
        created_at=datetime(2026, 5, 4, 12, 0, tzinfo=UTC),
    )
    session.add(s2)
    session.commit()

    # Both cases together: 1 with run, 1 without run -> rate is 50.00%
    all_rows = build_statistics_rows(session)
    assert len(all_rows) == 2
    c1_row = next(r for r in all_rows if r.inspection_case_id == str(c1.id))
    c2_row = next(r for r in all_rows if r.inspection_case_id == str(c2.id))

    assert c1_row.ocr_review_required is True
    assert c2_row.ocr_review_required is False

    data_both = calculate_quality_statistics_data(
        all_rows, "2026-05-01", "2026-05-31", datetime(2026, 5, 5, tzinfo=UTC)
    )
    assert data_both["ocr_review_rate"] == "50.00%"
