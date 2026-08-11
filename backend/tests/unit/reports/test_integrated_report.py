from __future__ import annotations

import io
from datetime import UTC, datetime

from openpyxl import load_workbook

from hyc_api.reports.deterministic import workbook_digest
from hyc_api.reports.integrated import render_integrated_inspection_report
from hyc_api.reports.sources import FrozenDecisionSource, LookedUpReferenceSource

_FROZEN = FrozenDecisionSource(
    payload={
        "overall_decision": "ACCEPTED",
        "spec_version": {"semantic_version": "1.2.0", "status": "ACTIVE"},
        "spec_items": [],
        "supplier_results": [],
        "internal_results": [],
        "item_decisions": [],
        "missing_policy": [],
        "sample_policy": [],
        "decision_reasons": {"final": "ACCEPTED", "reason": "ENGINE_MATCH"},
        "document_hashes": ["a" * 64],
        "approver": {"actor_id": "1", "role": "LEAD"},
        "lot_reference": {"lot_id": "lot-1"},
        "allocation_reference": {"allocation_id": "alloc-1"},
    },
    content_hash="b" * 64,
)


def _reference(material: str = "염화칼슘") -> LookedUpReferenceSource:
    return LookedUpReferenceSource(
        material_name=material,
        supplier_name="세계로비드",
        model_name="비드",
        documents=[("coa.pdf", "a" * 64)],
        nonconformances=[],
        attachments=[],
        observed_at=datetime(2026, 8, 11, 3, 0, tzinfo=UTC),
    )


def _sheets(payload: bytes) -> dict[str, list[list[object]]]:
    workbook = load_workbook(io.BytesIO(payload))
    return {
        name: [list(row) for row in workbook[name].iter_rows(values_only=True)]
        for name in workbook.sheetnames
    }


def test_every_prd_sheet_is_present() -> None:
    sheets = _sheets(
        render_integrated_inspection_report(_FROZEN, _reference(), include_audit=False)
    )
    assert list(sheets) == ["요약", "공급사결과", "자체검사", "판정근거", "부적합", "문서"]


def test_audit_sheet_appears_only_when_requested() -> None:
    with_audit = _sheets(
        render_integrated_inspection_report(_FROZEN, _reference(), include_audit=True)
    )
    assert "감사" in with_audit


def test_snapshot_sheets_carry_the_frozen_provenance_label() -> None:
    sheets = _sheets(
        render_integrated_inspection_report(_FROZEN, _reference(), include_audit=False)
    )
    assert sheets["판정근거"][0] == ["출처", "승인 시점 고정 (snapshot bbbbbbbbbbbb)"]


def test_lookup_sheets_carry_the_observed_at_label() -> None:
    sheets = _sheets(
        render_integrated_inspection_report(_FROZEN, _reference(), include_audit=False)
    )
    assert sheets["부적합"][0] == ["출처", "조회 시점 2026-08-11T03:00:00Z"]


def test_changing_a_lookup_does_not_change_snapshot_sheets() -> None:
    original = _sheets(
        render_integrated_inspection_report(_FROZEN, _reference(), include_audit=False)
    )
    renamed = _sheets(
        render_integrated_inspection_report(_FROZEN, _reference("다른 품목"), include_audit=False)
    )
    assert renamed["판정근거"] == original["판정근거"]
    assert renamed["공급사결과"] == original["공급사결과"]
    assert renamed["요약"] != original["요약"]


def test_identical_inputs_render_an_identical_digest() -> None:
    first = render_integrated_inspection_report(_FROZEN, _reference(), include_audit=False)
    second = render_integrated_inspection_report(_FROZEN, _reference(), include_audit=False)
    assert workbook_digest(first) == workbook_digest(second)
