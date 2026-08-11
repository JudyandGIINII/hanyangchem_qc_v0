from __future__ import annotations

import io
import zipfile

from hyc_api.reports.deterministic import SheetSpec, render_workbook, workbook_digest

_SHEETS = (
    SheetSpec(title="Summary", rows=[["항목", "값"], ["판정", "ACCEPTED"]]),
    SheetSpec(title="Items", rows=[["코드", "결과"], ["CA-01", "3.14"]]),
)


def test_same_input_renders_identical_bytes() -> None:
    first = render_workbook(_SHEETS)
    second = render_workbook(_SHEETS)
    assert workbook_digest(first) == workbook_digest(second)
    assert first == second


def test_zip_member_timestamps_are_pinned() -> None:
    # openpyxl writes the current clock into every zip member; an unpinned
    # timestamp makes the digest change once per second.
    with zipfile.ZipFile(io.BytesIO(render_workbook(_SHEETS))) as archive:
        stamps = {info.date_time for info in archive.infolist()}
    assert stamps == {(1980, 1, 1, 0, 0, 0)}


def test_changed_cell_changes_the_digest() -> None:
    # Guards against a render that pins so much it stops reflecting input.
    changed = (
        SheetSpec(title="Summary", rows=[["항목", "값"], ["판정", "REJECTED"]]),
        _SHEETS[1],
    )
    assert workbook_digest(render_workbook(_SHEETS)) != workbook_digest(render_workbook(changed))


def test_sheet_titles_and_cells_survive_the_round_trip() -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(render_workbook(_SHEETS)))
    assert workbook.sheetnames == ["Summary", "Items"]
    assert workbook["Items"].cell(row=2, column=1).value == "CA-01"
