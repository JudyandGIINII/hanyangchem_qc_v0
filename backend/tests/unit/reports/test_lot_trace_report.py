from __future__ import annotations

import io
import os
from collections.abc import Generator
from datetime import UTC, date, datetime
from decimal import Decimal
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from hyc_api.reports.deterministic import workbook_digest
from hyc_api.reports.lot_trace import render_lot_trace_report
from hyc_api.reports.sources import (
    FrozenDecisionSource,
    LookedUpReferenceSource,
    LotTraceAllocation,
    LotTraceCasePair,
    LotTraceSources,
    ReportSourceUnavailable,
    load_lot_trace_sources,
)
from hyc_data.models import (
    Base,
    InboundReceipt,
    Material,
    MaterialLot,
    ReceiptLotAllocation,
    Supplier,
)

POSTGRES_DSN = os.environ.get("HYC_P3_TEST_POSTGRES_DSN") or os.environ.get(
    "HYC_P2_TEST_POSTGRES_DSN"
)
# No postgres marker: the session fixture falls back to sqlite in-memory when no
# DSN is set, so these run in `make check` -- the gate that fires on every change.
# A module-level marker would deselect them there and they run in no postgres gate
# either, because those target integration/ rather than unit/.


@pytest.fixture
def session() -> Generator[Session, None, None]:
    if POSTGRES_DSN:
        engine = create_engine(POSTGRES_DSN)
    else:
        engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as s:
        yield s
    engine.dispose()


def _dummy_sources() -> LotTraceSources:
    alloc1 = LotTraceAllocation(
        allocation_id="alloc-uuid-1",
        inbound_no="INB-2026-001",
        receipt_date="2026-05-01",
        model_name="비드 25kg",
        quantity="50.00",
        quantity_unit="kg",
        status="CLOSED",
    )
    alloc2 = LotTraceAllocation(
        allocation_id="alloc-uuid-2",
        inbound_no="INB-2026-002",
        receipt_date="2026-05-15",
        model_name="비드 25kg",
        quantity="50.00",
        quantity_unit="kg",
        status="CLOSED",
    )
    frozen = FrozenDecisionSource(
        payload={
            "inspection_case_id": "case-uuid-1",
            "overall_decision": "ACCEPTED",
            "candidate_decision": "ACCEPTED",
            "spec_version": {"semantic_version": "1.0.0", "status": "ACTIVE"},
            "approver": {"actor_id": "lead-1", "role": "LEAD"},
            "allocation_reference": {"allocation_id": "alloc-uuid-1"},
        },
        content_hash="a" * 64,
    )
    ref = LookedUpReferenceSource(
        material_name="염화칼슘",
        supplier_name="세계로비드",
        model_name="비드 25kg",
        documents=[("coa.pdf", "b" * 64)],
        nonconformances=[
            {
                "ncr_number": "NCR-2026-001",
                "severity": "MAJOR",
                "description": "수분함량 초과",
                "status": "APPROVED",
            }
        ],
        attachments=["photo.jpg"],
        observed_at=datetime(2026, 8, 12, 0, 0, tzinfo=UTC),
    )
    return LotTraceSources(
        material_lot_id="lot-uuid-1",
        supplier_lot_no_raw="RAW-LOT-100",
        identity_key="LOT-CANONICAL-100",
        identity_status="CANONICAL",
        production_date_evidence="2026-04-15",
        merged_into_id="",
        supplier_name="세계로비드",
        material_name="염화칼슘",
        model_name="비드 25kg",
        allocations=[alloc1, alloc2],
        case_pairs=[LotTraceCasePair(frozen=frozen, reference=ref)],
        observed_at=datetime(2026, 8, 12, 0, 0, tzinfo=UTC),
    )


def test_lot_trace_sheet_layout_and_audit() -> None:
    sources = _dummy_sources()
    report_bytes = render_lot_trace_report(sources, include_audit=False)
    assert len(report_bytes) > 0
    assert workbook_digest(report_bytes) == workbook_digest(report_bytes)

    wb = load_workbook(io.BytesIO(report_bytes))
    assert list(wb.sheetnames) == [
        "LOT기본정보",
        "입고이력",
        "검사이력",
        "부적합",
        "문서",
        "생산및출하추적",
    ]

    report_with_audit = render_lot_trace_report(sources, include_audit=True)
    wb_audit = load_workbook(io.BytesIO(report_with_audit))
    assert "감사" in wb_audit.sheetnames


def test_lot_trace_downstream_erp_seam_is_empty() -> None:
    sources = _dummy_sources()
    report_bytes = render_lot_trace_report(sources, include_audit=False)
    wb = load_workbook(io.BytesIO(report_bytes))
    erp_sheet = wb["생산및출하추적"]
    rows = list(erp_sheet.iter_rows(values_only=True))

    assert rows[0][0] == "출처"
    assert "미연계" in str(rows[0][1])
    expected_header = (
        "투입 생산 LOT",
        "완제품 품목",
        "완제품 LOT",
        "투입 수량",
        "생산일자",
        "출하 LOT",
        "고객사",
    )
    assert rows[1] == expected_header
    assert len(rows) == 2


def test_lot_trace_all_cell_values_are_strings() -> None:
    sources = _dummy_sources()
    report_bytes = render_lot_trace_report(sources, include_audit=True)
    wb = load_workbook(io.BytesIO(report_bytes))
    for name in wb.sheetnames:
        sheet = wb[name]
        for row in sheet.iter_rows(values_only=True):
            for val in row:
                if val is not None:
                    err_msg = f"Cell in {name} is not str: {type(val)}"
                    assert isinstance(val, str), err_msg


def test_load_lot_trace_sources_not_found(session: Session) -> None:
    with pytest.raises(ReportSourceUnavailable) as exc_info:
        load_lot_trace_sources(session, uuid4())
    assert exc_info.value.code == "LOT_NOT_FOUND"


def test_lot_with_two_allocations_renders_exactly_two_inbound_data_rows(
    session: Session,
) -> None:
    """DECISIVE REGRESSION TEST: A lot with TWO allocations must render
    exactly two 입고이력 data rows.

    Proves the REP-003 defect (which previously rendered empty 입고이력 due to reading a missing
    'allocations' key from case snapshot) is fixed end-to-end against database sources.
    """
    supplier = Supplier(id=uuid4(), name="세계로비드", supplier_code="SUP-001")
    material = Material(id=uuid4(), name="염화칼슘", material_code="MAT-001")
    session.add_all([supplier, material])
    session.flush()

    lot = MaterialLot(
        id=uuid4(),
        supplier_id=supplier.id,
        material_id=material.id,
        identity_policy_version="v1",
        identity_key="LOT-SPLIT-001",
        identity_status="CANONICAL",
        supplier_lot_no_raw="RAW-SPLIT-100",
    )
    session.add(lot)
    session.flush()

    receipt1 = InboundReceipt(
        id=uuid4(),
        inbound_no="INB-2026-001",
        supplier_id=supplier.id,
        receipt_date=date(2026, 5, 1),
        status="CLOSED",
    )
    receipt2 = InboundReceipt(
        id=uuid4(),
        inbound_no="INB-2026-002",
        supplier_id=supplier.id,
        receipt_date=date(2026, 5, 15),
        status="CLOSED",
    )
    session.add_all([receipt1, receipt2])
    session.flush()

    alloc1 = ReceiptLotAllocation(
        id=uuid4(),
        inbound_receipt_id=receipt1.id,
        material_lot_id=lot.id,
        quantity=Decimal("50.00"),
        quantity_unit="kg",
    )
    alloc2 = ReceiptLotAllocation(
        id=uuid4(),
        inbound_receipt_id=receipt2.id,
        material_lot_id=lot.id,
        quantity=Decimal("50.00"),
        quantity_unit="kg",
    )
    session.add_all([alloc1, alloc2])
    session.commit()

    sources = load_lot_trace_sources(session, lot.id)
    assert len(sources.allocations) == 2

    report_bytes = render_lot_trace_report(sources)
    wb = load_workbook(io.BytesIO(report_bytes))
    alloc_sheet = wb["입고이력"]
    rows = list(alloc_sheet.iter_rows(values_only=True))

    data_rows = rows[2:]
    assert len(data_rows) == 2, f"Expected exactly 2 data rows, got {len(data_rows)}"

    assert data_rows[0][1] == "INB-2026-001"
    assert data_rows[1][1] == "INB-2026-002"
