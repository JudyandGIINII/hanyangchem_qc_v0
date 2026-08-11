from __future__ import annotations

import io
from datetime import UTC, datetime

from openpyxl import load_workbook

from hyc_api.reports.deterministic import workbook_digest
from hyc_api.reports.raw_data import render_raw_data_report
from hyc_api.reports.sources import FrozenDecisionSource, LookedUpReferenceSource

_FIXED_HEADERS = [
    "Model명",
    "품    명",
    "Lot Size",
    "검사구분",
    "입고일자",
    "검사일자",
    "검 사 자",
    "불량수량",
    "불 량 율",
    "제조업체",
    "Sample Size",
    "판   정",
    "처리방안",
]


def _reference() -> LookedUpReferenceSource:
    return LookedUpReferenceSource(
        material_name="염화칼슘",
        supplier_name="세계로비드",
        model_name="비드",
        documents=[("coa.pdf", "a" * 64)],
        nonconformances=[],
        attachments=[],
        observed_at=datetime(2026, 8, 11, 3, 0, tzinfo=UTC),
    )


def _frozen(item_count: int = 2) -> FrozenDecisionSource:
    spec_items = [
        {"id": f"item-{index}", "item_name": f"검사항목 {index}", "criterion": f">= {index}"}
        for index in range(1, item_count + 1)
    ]
    internal_results = [
        {
            "spec_item_id": f"item-{index}",
            "decision": "ACCEPTED",
            "samples": [
                {"sample_index": 1, "value": f"{index}.1", "decision": "ACCEPTED"},
                {"sample_index": 2, "value": f"{index}.2", "decision": "ACCEPTED"},
                {"sample_index": 3, "value": f"{index}.3", "decision": "ACCEPTED"},
            ],
        }
        for index in range(1, item_count + 1)
    ]
    return FrozenDecisionSource(
        payload={
            "overall_decision": "ACCEPTED",
            "spec_items": spec_items,
            "internal_results": internal_results,
            "approver": {"actor_id": 7, "role": "LEAD"},
        },
        content_hash="b" * 64,
    )


def _workbook(payload: bytes):
    return load_workbook(io.BytesIO(payload))


def test_raw_data_has_required_sheets_and_optional_audit() -> None:
    workbook = _workbook(render_raw_data_report(_frozen(), _reference(), include_audit=False))
    assert workbook.sheetnames == ["Raw_Data", "Measurements_Long", "Documents"]
    with_audit = _workbook(render_raw_data_report(_frozen(), _reference(), include_audit=True))
    assert with_audit.sheetnames == ["Raw_Data", "Measurements_Long", "Documents", "Audit"]


def test_raw_data_preserves_legacy_headers_and_complete_groups_beyond_column_56() -> None:
    worksheet = _workbook(render_raw_data_report(_frozen(item_count=10), _reference()))["Raw_Data"]
    headers = [cell.value for cell in worksheet[2]]
    assert headers[:13] == _FIXED_HEADERS
    assert worksheet.max_column == 63
    assert headers[-5:] == ["검사항목", "기준", "결과#19", "결과#20", "판정"]


def test_measurements_long_keeps_every_sample_and_all_rendered_cells_are_strings() -> None:
    workbook = _workbook(render_raw_data_report(_frozen(), _reference()))
    measurements = workbook["Measurements_Long"]
    rows = list(measurements.iter_rows(min_row=3, values_only=True))
    assert len(rows) == 6
    assert rows[-1] == ("검사항목 2", ">= 2", "3", "2.3", "ACCEPTED")
    assert all(
        isinstance(cell.value, str)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    )


def test_documents_and_provenance_follow_the_report_convention() -> None:
    workbook = _workbook(render_raw_data_report(_frozen(), _reference()))
    assert workbook["Raw_Data"]["A1"].value == "출처"
    assert workbook["Measurements_Long"]["B1"].value == "승인 시점 고정 (snapshot bbbbbbbbbbbb)"
    assert workbook["Documents"]["B1"].value == "조회 시점 2026-08-11T03:00:00Z"
    assert workbook["Documents"]["A3"].value == "coa.pdf"
    assert workbook["Documents"]["B3"].value == "a" * 64


def test_raw_data_is_byte_reproducible() -> None:
    first = render_raw_data_report(_frozen(), _reference())
    second = render_raw_data_report(_frozen(), _reference())
    assert workbook_digest(first) == workbook_digest(second)
