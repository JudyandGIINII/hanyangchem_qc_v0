from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

import hyc_api.master_import as master_import


def test_synthetic_workbook_builds_create_preview_for_every_entity() -> None:
    for entity in ("MATERIAL", "SUPPLIER", "MATERIAL_MODEL"):
        plan = master_import.build_master_import_plan(
            master_import.synthesize_master_import_workbook(
                entity, [("  A-001  ", "  Example name  "), (None, "No code is valid")]
            ),
            entity,
        )

        actual_rows = [
            (row.row_number, row.action, row.code, row.name, row.errors) for row in plan.rows
        ]
        assert actual_rows == [
            (2, "CREATE", "A-001", "Example name", ()),
            (3, "CREATE", None, "No code is valid", ()),
        ]
        assert (plan.created, plan.updated, plan.unchanged, plan.rejected) == (2, 0, 0, 0)


@pytest.mark.parametrize(
    "header",
    [
        ("code",),
        ("code", "name", "extra"),
        ("name", "code"),
    ],
)
def test_header_contract_rejects_the_entire_workbook(header: tuple[str, ...]) -> None:
    payload = master_import.synthesize_master_import_workbook("MATERIAL", [("M-1", "Material")])
    workbook = load_workbook(BytesIO(payload))
    worksheet = workbook.active
    if header == ("code",):
        worksheet.delete_cols(2)
    elif header == ("code", "name", "extra"):
        worksheet.insert_cols(3)
    for index, value in enumerate(header, start=1):
        worksheet.cell(row=1, column=index, value=value)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    plan = master_import.build_master_import_plan(buffer.getvalue(), "MATERIAL")

    assert plan.created == plan.updated == plan.unchanged == 0
    assert plan.rejected == 1
    assert len(plan.rows) == 1
    assert plan.rows[0].row_number == 1
    assert plan.rows[0].action == "REJECT"
    assert "header must contain exactly the columns 'code', 'name' in that order" in (
        plan.rows[0].errors[0]
    )


def test_later_duplicate_non_null_code_is_rejected_but_blank_codes_are_not() -> None:
    plan = master_import.build_master_import_plan(
        master_import.synthesize_master_import_workbook(
            "SUPPLIER",
            [
                (None, "First no-code supplier"),
                ("", "Second no-code supplier"),
                (" S-1 ", "First"),
                ("S-1", "Later"),
            ],
        ),
        "SUPPLIER",
    )

    assert [row.action for row in plan.rows] == ["CREATE", "CREATE", "CREATE", "REJECT"]
    assert plan.rows[-1].errors == ("duplicate non-null code within workbook: 'S-1'",)
    assert (plan.created, plan.updated, plan.unchanged, plan.rejected) == (3, 0, 0, 1)


def test_blank_name_is_rejected_with_a_preview_reason() -> None:
    plan = master_import.build_master_import_plan(
        master_import.synthesize_master_import_workbook("MATERIAL_MODEL", [("MM-1", "  ")]),
        "MATERIAL_MODEL",
    )

    assert plan.rows[0].action == "REJECT"
    assert plan.rows[0].errors == ("name is required",)
    assert plan.rejected == 1


def test_invalid_entity_is_rejected() -> None:
    with pytest.raises(ValueError, match="unsupported master import entity"):
        master_import.synthesize_master_import_workbook("UNKNOWN", [])  # type: ignore[arg-type]


def test_module_exposes_no_apply_style_function() -> None:
    prohibited = {"apply", "commit", "save", "write"}
    assert prohibited.isdisjoint(vars(master_import))
