from __future__ import annotations

import io
from datetime import datetime
from uuid import UUID

import pytest
from openpyxl import load_workbook

from hyc_api.reports.statistics import (
    build_statistics_rows,
    render_supplier_quality_statistics_report,
)

pytestmark = pytest.mark.postgres


def _approved_case_id(p3) -> UUID:
    flow = p3.reviewed()
    inspection_id = flow["inspection_id"]
    ready = p3.clear_hold(inspection_id)
    submitted = p3.submit(inspection_id, ready["version"])
    approved = p3.approve(inspection_id, submitted["version"])
    assert approved.status_code == 200, approved.text
    return UUID(inspection_id)


def test_statistics_endpoint_requires_auth(p3) -> None:
    res = p3.client.get("/api/v1/statistics/quality?period_start=2026-05-01&period_end=2026-05-31")
    assert res.status_code == 401


def test_statistics_endpoint_inverted_period_returns_422(p3) -> None:
    res = p3.client.get(
        "/api/v1/statistics/quality?period_start=2026-06-01&period_end=2026-05-01",
        headers=p3.lead,
    )
    assert res.status_code == 422
    assert "period_start" in res.text


def test_statistics_endpoint_and_workbook_cross_check(p3, p3_engine_storage) -> None:
    """CRITICAL REGRESSION TEST: Cross-check JSON API response against Excel workbook output.

    Feeds ONE identical set of rows through both the GET /api/v1/statistics/quality
    endpoint and render_supplier_quality_statistics_report, then asserts all numbers match.
    """
    _case_id = _approved_case_id(p3)

    # 1. Fetch JSON response from API endpoint
    res = p3.client.get(
        "/api/v1/statistics/quality?period_start=2026-08-01&period_end=2026-08-31",
        headers=p3.lead,
    )
    assert res.status_code == 200, res.text
    json_resp = res.json()

    assert json_resp["population"]["approved_case_count"] >= 1

    # 2. Render Excel report with exact same DB rows
    with p3_engine_storage.session_factory() as session:
        db_rows = build_statistics_rows(session)
        observed_at = datetime.fromisoformat(json_resp["observed_at"])
        wb_bytes = render_supplier_quality_statistics_report(
            db_rows,
            period_start="2026-08-01",
            period_end="2026-08-31",
            observed_at=observed_at,
        )

    # 3. Assert numbers in Excel workbook equal numbers in JSON response
    wb = load_workbook(io.BytesIO(wb_bytes))
    summary_rows = list(wb["품질통계요약"].iter_rows(values_only=True))
    summary_map = {str(r[0]): str(r[1]) for r in summary_rows[1:] if len(r) >= 2}

    assert summary_map["총 입고/검사 건수"] == str(json_resp["population"]["approved_case_count"])
    assert summary_map["평균 처리기간 (일)"] == json_resp["average_handling_days"]
    assert summary_map["OCR 검토 필요율"] == json_resp["ocr_review_rate"]
