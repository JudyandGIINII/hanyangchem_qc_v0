"""Pure workbook parsing and preview planning for P6 master imports.

Lives in ``hyc_api`` rather than ``hyc_domain`` because it parses xlsx: the
domain layer is guarded against infrastructure imports, and openpyxl is one.
The rules it applies are still pure -- it touches no Session and no file.
"""

from __future__ import annotations

import io
from collections.abc import Sequence
from dataclasses import dataclass
from typing import Literal

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

type MasterImportEntity = Literal["MATERIAL", "SUPPLIER", "MATERIAL_MODEL"]
type ImportAction = Literal["CREATE", "UPDATE", "UNCHANGED", "REJECT"]

_ENTITIES: frozenset[str] = frozenset({"MATERIAL", "SUPPLIER", "MATERIAL_MODEL"})
_EXPECTED_HEADER = ("code", "name")


@dataclass(frozen=True, slots=True)
class ImportRowPlan:
    row_number: int
    entity: MasterImportEntity
    action: ImportAction
    code: str | None
    name: str
    errors: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class ImportPlan:
    rows: tuple[ImportRowPlan, ...]
    created: int
    updated: int
    unchanged: int
    rejected: int


def _validate_entity(entity: str) -> MasterImportEntity:
    if entity not in _ENTITIES:
        raise ValueError(f"unsupported master import entity: {entity!r}")
    return entity  # type: ignore[return-value]


def _cell_text(value: object | None) -> str:
    return "" if value is None else str(value).strip()


def _count_actions(rows: tuple[ImportRowPlan, ...]) -> tuple[int, int, int, int]:
    return (
        sum(row.action == "CREATE" for row in rows),
        sum(row.action == "UPDATE" for row in rows),
        sum(row.action == "UNCHANGED" for row in rows),
        sum(row.action == "REJECT" for row in rows),
    )


def _header_rejection(entity: MasterImportEntity, header: tuple[str, ...]) -> ImportPlan:
    row = ImportRowPlan(
        row_number=1,
        entity=entity,
        action="REJECT",
        code=None,
        name="",
        errors=(
            "header must contain exactly the columns 'code', 'name' in that order; "
            f"received {header!r}",
        ),
    )
    return ImportPlan(rows=(row,), created=0, updated=0, unchanged=0, rejected=1)


def build_master_import_plan(workbook_bytes: bytes, entity: MasterImportEntity) -> ImportPlan:
    """Parse a synthetic or caller-provided workbook into a no-side-effect preview.

    This parsing-only half has no master-state input, so each valid row is a
    ``CREATE`` candidate. The database-layer owner later reconciles previewed
    values with current state before presenting UPDATE or UNCHANGED outcomes.
    """

    validated_entity = _validate_entity(entity)
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        row_values = worksheet.iter_rows(values_only=True)
        header_values = next(row_values, ())
        header = tuple(_cell_text(value) for value in header_values)
        if header != _EXPECTED_HEADER:
            return _header_rejection(validated_entity, header)

        seen_codes: set[str] = set()
        rows: list[ImportRowPlan] = []
        for row_number, row in enumerate(row_values, start=2):
            code_text = _cell_text(row[0] if len(row) > 0 else None)
            name = _cell_text(row[1] if len(row) > 1 else None)
            code = code_text or None
            errors: list[str] = []
            if not name:
                errors.append("name is required")
            if code is not None and code in seen_codes:
                errors.append(f"duplicate non-null code within workbook: {code!r}")
            if code is not None:
                seen_codes.add(code)
            rows.append(
                ImportRowPlan(
                    row_number=row_number,
                    entity=validated_entity,
                    action="REJECT" if errors else "CREATE",
                    code=code,
                    name=name,
                    errors=tuple(errors),
                )
            )
        planned_rows = tuple(rows)
        created, updated, unchanged, rejected = _count_actions(planned_rows)
        return ImportPlan(
            rows=planned_rows,
            created=created,
            updated=updated,
            unchanged=unchanged,
            rejected=rejected,
        )
    finally:
        workbook.close()


def synthesize_master_import_workbook(
    entity: MasterImportEntity, rows: Sequence[tuple[str | None, str]]
) -> bytes:
    """Build an in-memory, synthetic sample workbook using the import layout."""

    _validate_entity(entity)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = entity
    worksheet.append(list(_EXPECTED_HEADER))
    for code, name in rows:
        worksheet.append([code, name])
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
